#!/usr/bin/env python3
"""Build a Review Set shaping workbook (.xlsx) from a tab-separated plan file.

    python3 build_review_set_workbook.py <input.tsv> <output.xlsx> "<Review Set title>" "<description>"

Requires openpyxl. The repo's Python is externally managed, so use a venv:
    python3 -m venv /tmp/xlsxvenv && /tmp/xlsxvenv/bin/pip install openpyxl
    /tmp/xlsxvenv/bin/python docs/curriculum/build_review_set_workbook.py ...

INPUT COLUMNS (tab-separated, header row required, order irrelevant):
    plan_no            integer, 1-based; groups rows into Subject Plan sheets
    subject_plan       plan title, emoji allowed; must be constant per plan_no
    plan_description   one sentence; must be constant per plan_no; printed ONCE per sheet
    section            section name; repeats across its rows, printed once per group
    note_title         the note
    note_subject       canonical Subject metadata (NOT the section name)
    domain_context     enum value, or "(unset)"
    status             Existing | Reuse | New | Excluded

Row ORDER is authoritative: sections appear in first-seen order, notes in file order.
Nothing is sorted, so the strategist's sequencing survives into the workbook.

See docs/curriculum/review-set-workbook-spec.md for why each sheet exists.
"""
import sys, csv, collections
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

STATUS_FILL = {"Existing":"D5E8D4","Reuse":"DAE8FC","New":"FFF2CC","Excluded":"F8CECC"}
DC_FILL = {"(unset)":"EAEAEA","ENGINEERING_SCIENCES":"E1D5E7","ENGINEERING_MATHEMATICS":"D5E8D4",
           "CIVIL_ENGINEERING":"FFE6CC","PROFESSIONAL_PRACTICE_AND_REGULATION":"DAE8FC",
           "ENGINEERING_MATHEMATICS ":"D5E8D4"}
HDR = PatternFill("solid", fgColor="44546A")
_T = Side(style="thin", color="D0D0D0")
BOX = Border(left=_T, right=_T, top=_T, bottom=_T)
EXCLUDED = "Excluded"


def _head(ws, row, labels):
    for i, h in enumerate(labels, 1):
        c = ws.cell(row=row, column=i, value=h)
        c.font = Font(bold=True, color="FFFFFF"); c.fill = HDR; c.border = BOX


def _banner(ws, cell, text, span, height, **font):
    ws[cell] = text
    ws[cell].alignment = Alignment(wrap_text=True, vertical="top")
    if font: ws[cell].font = Font(**font)
    ws.merge_cells(span); ws.row_dimensions[int(cell[1:])].height = height


def _widths(ws, widths):
    for col, w in zip("ABCDEFGH", widths):
        ws.column_dimensions[col].width = w


def sheet_name(plan_no, title):
    """Excel caps sheet names at 31 chars and forbids : \\ / ? * [ ]."""
    clean = "".join(ch for ch in title if ch not in ':\\/?*[]').strip()
    clean = "".join(ch for ch in clean if ch.isascii()).strip()
    return f"{plan_no} {clean}"[:31]


def build(rows, out, set_title, set_desc):
    plans = collections.OrderedDict()
    for r in rows:
        plans.setdefault(int(r["plan_no"]), {"title": r["subject_plan"],
                                             "desc": r["plan_description"],
                                             "sections": collections.OrderedDict()})
        plans[int(r["plan_no"])]["sections"].setdefault(r["section"], []).append(r)

    # a note carried by two plans is ONE canonical note, not a duplicate
    shared = collections.Counter()
    for r in rows:
        if r["status"] != EXCLUDED:
            shared[r["note_title"]] += 1

    wb = Workbook(); ov = wb.active; ov.title = "Overview"
    ov["A1"] = f"{set_title} — target shape"; ov["A1"].font = Font(size=15, bold=True)
    _banner(ov, "A2", set_desc, "A2:H2", 42)
    _banner(ov, "A4", "Status: Existing = already in the set · Reuse = exists elsewhere, add it · "
                      "New = needs authoring · Excluded = deliberately held out.  "
                      "See the Domain Context sheet for the two hard rules on that column.",
            "A4:H4", 16, italic=True, color="B06000")
    _head(ov, 6, ["#", "Subject Plan", "Sections", "In set", "Existing", "Reuse", "New", "Excluded"])
    r, tot = 7, collections.Counter()
    for pno, p in plans.items():
        st = collections.Counter(n["status"] for ns in p["sections"].values() for n in ns)
        tot.update(st)
        vals = [pno, p["title"], len(p["sections"]),
                sum(v for k, v in st.items() if k != EXCLUDED),
                st["Existing"], st["Reuse"], st["New"], st[EXCLUDED]]
        for i, v in enumerate(vals, 1):
            ov.cell(row=r, column=i, value=v).border = BOX
        r += 1
    for i, v in enumerate(["", "TOTAL", "", sum(v for k, v in tot.items() if k != EXCLUDED),
                           tot["Existing"], tot["Reuse"], tot["New"], tot[EXCLUDED]], 1):
        c = ov.cell(row=r, column=i, value=v); c.font = Font(bold=True); c.border = BOX
    _widths(ov, [5, 42, 10, 10, 10, 9, 8, 10]); ov.freeze_panes = "A7"

    dc = wb.create_sheet("Domain Context")
    dc["A1"] = "Domain Context by Subject"; dc["A1"].font = Font(size=14, bold=True)
    _banner(dc, "A2", "RULE 1 — (unset) is only legal on a note with ONE Applicable Program. "
            "NoteApplicableProgramsService rejects a save with 2+ programs and no Domain Context "
            "(MultiProgramDomainContextRequiredException).", "A2:C2", 44, color="B06000", bold=True)
    _banner(dc, "A3", "RULE 2 — Existing and Reuse notes ALREADY have a Domain Context. The value here is a "
            "recommendation, not a blank field; changing it is a separate decision from placing the note, and "
            "it affects FUTURE generation only.", "A3:C3", 44, color="B06000", bold=True)
    _banner(dc, "A4", "NOTE — (unset) falls back to the program name. If that name matches no quantitative "
            "keyword, computation guidance stays OFF for every quiz generated from the note.", "A4:C4", 30,
            italic=True, color="666666")
    _head(dc, 6, ["Note subject", "Domain Context", "Notes in set"])
    per = collections.OrderedDict()
    for r_ in rows:
        per.setdefault((r_["note_subject"], r_["domain_context"]), 0)
        per[(r_["note_subject"], r_["domain_context"])] += 1
    r = 7
    for (subj, v), n in sorted(per.items()):
        for i, x in enumerate([subj, v, n], 1):
            c = dc.cell(row=r, column=i, value=x); c.border = BOX
            if i == 2:
                c.fill = PatternFill("solid", fgColor=DC_FILL.get(v, "FFFFFF")); c.font = Font(bold=True)
        r += 1
    _widths(dc, [34, 40, 14]); dc.freeze_panes = "A7"

    for pno, p in plans.items():
        ws = wb.create_sheet(sheet_name(pno, p["title"]))
        ws["A1"] = p["title"]; ws["A1"].font = Font(size=14, bold=True); ws.merge_cells("A1:G1")
        _banner(ws, "A2", p["desc"], "A2:G2", 42)
        _head(ws, 4, ["Section", "#", "Note title", "Note subject", "Domain Context", "Status", "Flags"])
        r = 5
        for sec, notes in p["sections"].items():
            for i, n in enumerate(notes):
                flags = []
                if shared[n["note_title"]] > 1:
                    flags.append("same canonical note in another plan — do not duplicate")
                if n["status"] in ("Reuse", "Existing"):
                    flags.append("context already set — verify before changing")
                elif n["domain_context"] == "(unset)":
                    flags.append("unset requires a SINGLE applicable program")
                vals = [sec if i == 0 else "", i + 1, n["note_title"], n["note_subject"],
                        n["domain_context"], n["status"], " · ".join(flags)]
                for ci, x in enumerate(vals, 1):
                    c = ws.cell(row=r, column=ci, value=x); c.border = BOX
                    if ci == 1: c.font = Font(bold=True)
                    if ci == 2: c.alignment = Alignment(horizontal="center")
                    if ci == 5: c.fill = PatternFill("solid", fgColor=DC_FILL.get(x, "FFFFFF"))
                    if ci == 6 and x in STATUS_FILL: c.fill = PatternFill("solid", fgColor=STATUS_FILL[x])
                    if ci == 7 and flags: c.font = Font(italic=True, color="B06000", size=9)
                r += 1
            r += 1
        _widths(ws, [34, 4, 60, 30, 34, 11, 58]); ws.freeze_panes = "A5"

    bs = wb.create_sheet("By Subject (bulk generate)")
    bs["A1"] = "Generation batches — Bulk Generate applies ONE subject and ONE Domain Context per batch"
    bs["A1"].font = Font(size=13, bold=True); bs.merge_cells("A1:F1")
    _banner(bs, "A2", "Only 'New' notes appear. Each block is one Bulk Generate run: set the Subject and Domain "
            "Context shown, paste the titles as topics.", "A2:F2", 16, italic=True, color="666666")
    _head(bs, 4, ["Note subject", "Domain Context", "New notes", "Subject Plan", "Section", "Note title"])
    by = collections.defaultdict(list)
    for pno, p in plans.items():
        for sec, notes in p["sections"].items():
            for n in notes:
                if n["status"] == "New":
                    by[(n["note_subject"], n["domain_context"])].append((p["title"], sec, n["note_title"]))
    r = 5
    for key in sorted(by, key=lambda k: (-len(by[k]), k[0])):
        subj, v = key
        for i, (pl, sec, t) in enumerate(by[key]):
            vals = [subj if i == 0 else "", v if i == 0 else "", len(by[key]) if i == 0 else "", pl, sec, t]
            for ci, x in enumerate(vals, 1):
                c = bs.cell(row=r, column=ci, value=x); c.border = BOX
                if ci in (1, 3): c.font = Font(bold=True)
                if ci == 2 and i == 0:
                    c.fill = PatternFill("solid", fgColor=DC_FILL.get(v, "FFFFFF")); c.font = Font(bold=True)
            r += 1
        r += 1
    _widths(bs, [32, 34, 10, 36, 32, 60]); bs.freeze_panes = "A5"

    wb.save(out)
    return plans, tot, by


def main():
    if len(sys.argv) != 5:
        print(__doc__); sys.exit(1)
    src, out, title, desc = sys.argv[1:5]
    with open(src, encoding="utf-8-sig", newline="") as f:
        rows = list(csv.DictReader(f, delimiter="\t"))
    required = {"plan_no","subject_plan","plan_description","section","note_title",
                "note_subject","domain_context","status"}
    missing = required - set(rows[0])
    if missing:
        sys.exit(f"input is missing required columns: {sorted(missing)}")
    bad = {r["status"] for r in rows} - set(STATUS_FILL)
    if bad:
        sys.exit(f"unknown status values: {sorted(bad)} (allowed: {sorted(STATUS_FILL)})")
    plans, tot, by = build(rows, out, title, desc)
    print(f"saved {out}")
    print(f"{len(plans)} plans · {len(rows)} rows · "
          + " · ".join(f"{k}={v}" for k, v in sorted(tot.items())))
    print(f"{len(by)} bulk-generate batches")


if __name__ == "__main__":
    main()
