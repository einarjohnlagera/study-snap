#!/usr/bin/env python3
"""Build the strategist INPUT workbook — one sheet per query result.

    python3 build_strategist_inputs_workbook.py <results_dir> <output.xlsx> "<Review Set name>"

`results_dir` holds one CSV per query, named q0.csv … q7.csv. Export them straight from the
SQL client; a header row is expected. Missing files are skipped, so you can build the workbook
from whichever queries you have run.

WHY THIS EXISTS. The alternative is pasting seven raw CSV blobs into a chat window, where a
long string_agg column (Q4's `titles`) becomes an unreadable wall and column alignment is lost.
One sheet per query keeps each result legible, and the header on each sheet tells the reader
what the query answers and whether it is meant for the strategist at all — Q7 is not.

Requires openpyxl:
    python3 -m venv /tmp/xlsxvenv && /tmp/xlsxvenv/bin/pip install openpyxl
"""
import sys, csv, os
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

# sheet name, what it answers, whether to hand it to the strategist
QUERIES = {
 "q0": ("Q0 Review sets", "Every root Review Set with child-plan and note counts.",
        "HAND OVER — the two relevant rows. This is the size gap being closed."),
 "q1": ("Q1 Target today (notes)", "The target set as it stands, note by note, with section and position.",
        "HAND OVER if under ~200 rows — it is the only place missing sections are visible."),
 "q2": ("Q2 Target today (skeleton)", "The target set as plan → section → counts.",
        "HAND OVER — always."),
 "q3": ("Q3 Benchmark shape", "A comprehensive set at summary level.",
        "HAND OVER — this is what 'comprehensive' means concretely."),
 "q4": ("Q4 Ready-to-add pool", "Notes already tagged for the target program but NOT in the set.",
        "HAND OVER — the cheapest coverage available. ⚠️ A CANDIDATE POOL, NOT AN ANSWER: "
        "program tags were partly set by a profile default, so expect material that shares a "
        "jobsite with the discipline without belonging in its licensure review."),
 "q5": ("Q5 Overlap map", "Per benchmark subject: notes, how many are target-tagged, how many already in the set.",
        "HAND OVER — a high count with zero tagged is a metadata decision, not an authoring job."),
 "q6": ("Q6 Catalog programs", "Exact catalog program names and how many notes carry each.",
        "HAND OVER trimmed to programs with real counts, so proposals name programs that exist."),
 "q7": ("Q7 Reconciliation", "Existing notes in the set grouped by Subject.",
        "⚠️ DO NOT HAND OVER. This is for reconciling a returned proposal's `status` column "
        "afterwards. Match on KNOWLEDGE, not string equality."),
}
HDR = PatternFill("solid", fgColor="44546A")
_T = Side(style="thin", color="D0D0D0")
BOX = Border(left=_T, right=_T, top=_T, bottom=_T)


def add_sheet(wb, key, rows):
    title, answers, handover = QUERIES[key]
    ws = wb.create_sheet(title[:31])
    ncols = max(len(r) for r in rows) if rows else 1
    span = f"A1:{chr(64 + min(ncols, 26))}1"
    ws["A1"] = answers; ws["A1"].font = Font(size=12, bold=True); ws.merge_cells(span)
    ws["A2"] = handover
    ws["A2"].font = Font(italic=True, color="B06000" if "DO NOT" in handover or "⚠️" in handover else "666666")
    ws["A2"].alignment = Alignment(wrap_text=True, vertical="top")
    ws.merge_cells(span.replace("1", "2")); ws.row_dimensions[2].height = 30
    for ci, h in enumerate(rows[0], 1):
        c = ws.cell(row=4, column=ci, value=h)
        c.font = Font(bold=True, color="FFFFFF"); c.fill = HDR; c.border = BOX
    for ri, row in enumerate(rows[1:], 5):
        for ci, v in enumerate(row, 1):
            c = ws.cell(row=ri, column=ci, value=v); c.border = BOX
            c.alignment = Alignment(vertical="top", wrap_text=len(str(v)) > 80)
    for ci in range(1, ncols + 1):
        width = max((len(str(r[ci - 1])) for r in rows if ci <= len(r)), default=10)
        ws.column_dimensions[chr(64 + ci)].width = min(max(width + 2, 10), 70)
    ws.freeze_panes = "A5"
    return len(rows) - 1


def main():
    if len(sys.argv) != 4:
        print(__doc__); sys.exit(1)
    src, out, name = sys.argv[1:4]
    wb = Workbook(); ov = wb.active; ov.title = "README"
    ov["A1"] = f"{name} — strategist inputs"; ov["A1"].font = Font(size=15, bold=True)
    ov["A2"] = ("One sheet per query from docs/curriculum/review-set-reshape-read.sql. Each sheet's header "
                "says what the query answers and whether to hand it to the strategist.")
    ov["A2"].alignment = Alignment(wrap_text=True, vertical="top"); ov.merge_cells("A2:C2")
    ov.row_dimensions[2].height = 30
    for ci, h in enumerate(["Sheet", "Rows", "Hand to strategist?"], 1):
        c = ov.cell(row=4, column=ci, value=h)
        c.font = Font(bold=True, color="FFFFFF"); c.fill = HDR; c.border = BOX
    r, built = 5, 0
    for key in sorted(QUERIES):
        path = os.path.join(src, f"{key}.csv")
        if not os.path.exists(path):
            continue
        with open(path, encoding="utf-8-sig", newline="") as f:
            rows = [row for row in csv.reader(f) if any(x.strip() for x in row)]
        if not rows:
            continue
        n = add_sheet(wb, key, rows); built += 1
        for ci, v in enumerate([QUERIES[key][0], n, "no — reconciliation" if key == "q7" else "yes"], 1):
            ov.cell(row=r, column=ci, value=v).border = BOX
        r += 1
    if not built:
        sys.exit(f"no q*.csv files found in {src}")
    for col, w in zip("ABC", [28, 8, 24]):
        ov.column_dimensions[col].width = w
    wb.save(out)
    print(f"saved {out} — {built} query sheets")


if __name__ == "__main__":
    main()
